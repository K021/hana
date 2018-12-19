import sys
import time
from datetime import datetime

from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QAxContainer import *
from openpyxl import Workbook

from functions import encrypt, decrypt
from config_secret import ENCRYPTION_KEY


app = QApplication(sys.argv)


class Hana(QAxWidget):
    _PROG_ID = 'HFCOMMAGENT.HFCommAgentCtrl.1'  # 하나대투 1Q OpenApi Program ID

    def __init__(self, user_id, user_pw, cert_pw):
        super().__init__()
        # 이 과정이 실행되기 위해선 관리자 권한이 필요할 수도 있다.

        self.setControl(self.prog_id)  # 이 과정이 실행되기 위해선 관리자 권한이 필요하다.
        self._user_id = encrypt(ENCRYPTION_KEY, user_id)
        self._user_pw = encrypt(ENCRYPTION_KEY, user_pw)
        self._cert_pw = encrypt(ENCRYPTION_KEY, cert_pw)

        self.OnGetFidData.connect(self.process_event_fid_data)
        self.OnGetRealData.connect(self.process_event_real_data)

        self.fid = '2'
        self.tig_data = list()

        self.event_connect_loop = QEventLoop()

    def tig_data_dump(self, file_name=None):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'tig_data'
        if not file_name:
            file_name = f'tig_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        sheet1.cell(row=1, column=1).value = '시간'
        sheet1.cell(row=1, column=2).value = '가격'
        for i, tig in enumerate(self.tig_data):
            sheet1.cell(row=i+2, column=1).value = tig[0]
            sheet1.cell(row=i+2, column=2).value = tig[1]

        wb.save(filename=file_name)

    @staticmethod
    def _block_to_tig_list(block, futures=False):
        """
        tig data 를 담고 있는 block 처리 함수
        Example:
            - block: '\x1e20181217\x1e151149\x1e11410\x1d\x1e20181217\x1e151133\x1e11415\x1f'
            - return:
                tigs = [
                    [datetime.datetime(2018, 12, 17, 15, 19, 57), 11385],
                    [datetime.datetime(2018, 12, 17, 15, 19, 56), 11390],
                ]
        :param block: tig data 를 담고 있는 문자열로, request_fid() 또는 request_fid_array() 호출시 발생하는 이벤트에서 반환된다.
        :return: tig 정보 list
        """
        # 각 틱은 '\x1d' 를 기준으로 구분된다.
        # 한 틱 내 값과 날짜는 '\x1e'를 기준으로 구분된다.
        # block 문자열 마지막의 '\x1f' 를 제거해준다.
        # block 예시: '\x1e20181217\x1e151149\x1e11410\x1d\x1e20181217\x1e151133\x1e11415\x1f'
        tigs = block[:-1].split('\x1d') if block[-1] == '\x1f' else block.split('\x1d')
        # tigs = [
        #     '\x1e20181217\x1e151957\x1e11385',
        #     '\x1e20181217\x1e151956\x1e11390'
        # ]
        tigs = [tig.split('\x1e') for tig in tigs]
        # tigs = [
        #     ['20181217', '151957', '11385'],
        #     ['20181217', '151956', '11390']
        # ]

        # 선물의 경우, 종가 값을 시간이 888888인 틱으로 나타내므로, 해당 틱 제거
        if futures:
            tigs = filter(lambda x: '888888' not in x[1], tigs)

        # 일자와 시간 결합
        tigs = [[ti[0] + ti[1], ti[2]] for ti in tigs]
        # tigs = [
        #     ['20181217151957', '11385'],
        #     ['20181217151956', '11390']
        # ]
        if not futures:  # 주식일 경우 가격 정수화
            tigs = [[datetime.strptime(ti[0], '%Y%m%d%H%M%S'), int(ti[1])] for ti in tigs]
        else:  # 선물일 경우 가격 실수화
            tigs = [[datetime.strptime(ti[0], '%Y%m%d%H%M%S'), float(ti[1])] for ti in tigs]
        # tigs = [
        #     [datetime.datetime(2018, 12, 17, 15, 19, 57), 11385],
        #     [datetime.datetime(2018, 12, 17, 15, 19, 56), 11390]
        # ]
        return tigs

    def _tig_validation(self, tigs, today=True):
        """
        tig list 를 받아서, 그중 유효한 값의 인덱스를 반환하는 함수
        :param tigs: tig list
        :param today: 오늘의 틱 데이터만 가져올 것인지를 결정하는 값
        :return: tigs 안의 유효한 tig 의 인덱스. 유효값이 없는 경우는 None
        """
        # 틱을 처음 가져올 때
        if not self.tig_data:
            # 오늘의 틱만 가져오기 위한 과정
            if today:
                dt_today = datetime.today()
                dt_today = dt_today.replace(hour=0, minute=0, second=0, microsecond=0)
                idx = 0
                for index, tig in enumerate(tigs):
                    if tig[0] < dt_today:  # 오늘이 아닌 값 제외
                        continue
                    else:
                        idx = index  # 오늘인 값 인덱스 저장하고 반복문 나가기
                        break
                tigs = tigs[idx:]
                if tigs:  # 인덱스가 tig 길이를 초과하지 않을때 (유효한 인덱스일 때)
                    return idx
                return None

            return 0  # today == False 인 경우, 첫번째 인덱스 반환

        last_time = self.tig_data[-1][0]  # 지금 까지 수집한 틱 데이터 중 가장 마지막 값의 시각
        last_time_count = 0  # 그 시간의 틱이 몇개인가
        for tig in reversed(self.tig_data):
            if tig[0] == last_time:
                last_time_count += 1
            else:
                break
        for index, tig in enumerate(tigs):
            if tig[0] < last_time:  # 겹치는 것 버림
                continue
            # 틱 데이터 중, 시간이 중복되는 틱이 있다.
            # 그것을 고려하여 중복되는 값을 버린다.
            elif tig[0] == last_time:
                if last_time_count == 0:
                    return index  # 중복되지 않는 틱의 인덱스값 반환
                else:
                    last_time_count -= 1
            elif tig[0] > last_time:
                return index  # 중복되지 않는 틱의 인덱스값 반환

    def process_event_fid_data(self, request_id, block, block_len):
        is_futures = True if '.' in block else False  # 가져오는 값이 선물인지 주가인지
        tig_infos = self._block_to_tig_list(block, futures=is_futures)  # block 을 tig info list 로 변환
        tig_infos.sort(key=lambda x: x[0])  # 시간 오름차순 순서대로 정렬
        index = self._tig_validation(tig_infos)  # 유효한 tig 의 인덱스 값 추출
        if index is not None:  # 유효한 값이 존재하는 경우 self.tig_data 에 추가
            self.tig_data.extend(tig_infos[index:])

        print()
        print('==== Request Information ====')
        print('1. Request ID:', request_id)
        print('2. Block:', block)
        if index is not None:
            print('3. Tig_data: \n', tig_infos[index:])
        else:
            print('3. Tig_data: 유효한 틱 데이터가 없습니다.')

        self.event_connect_loop.exit()

    def process_event_real_data(self, real_name, real_key, block, block_len):
        print('Real Name', real_name)
        print('Real Key', real_key)
        print('Block:\n', block)
        print('Block Length', block_len)
        self.event_connect_loop.exit()

    # 테스트 코드
    def get_fid_data_test(self, stock_code='229200', market_code='J', gid='1000'):
        rid = self.creat_request_id()
        assert rid, 'Request ID 가 생성되지 않았습니다.'
        self.set_fid_input_data(rid, '9002', stock_code)  # 종목 코드
        self.set_fid_input_data(rid, '9001', market_code)  # 주식시장
        self.set_fid_input_data(rid, 'GID', gid)
        self.request_fid(rid, '1,3,9,8,13,14,15,10,4,266,267', '9999')

    def get_fid_tig_data_test(self, stock_code='229200', market_code='J', gid='1000', day1='20181213', day2='20181213'):
        rid = self.creat_request_id()
        assert rid, 'Request ID 가 생성되지 않았습니다.'
        self.set_fid_input_data(rid, '9001', market_code)
        self.set_fid_input_data(rid, '9002', stock_code)
        self.set_fid_input_data(rid, '9008', '4')
        self.set_fid_input_data(rid, '9034', day1)
        self.set_fid_input_data(rid, '9035', day2)
        self.set_fid_input_data(rid, 'GID', gid)
        self.request_fid(rid, '1,3,9,8,4,13,267', '9999')

    def get_fid_tig_data_test_futures(self, stock_code='101P3000', market_code='F', gid='1002', day1=''):
        rid = self.creat_request_id()
        assert rid, 'Request ID 가 생성되지 않았습니다.'
        self.set_fid_input_data(rid, '9001', market_code)
        self.set_fid_input_data(rid, '9002', stock_code)
        self.set_fid_input_data(rid, '9034', day1)
        # self.set_fid_input_data(rid, '9035', day2)
        self.set_fid_input_data(rid, 'GID', gid)
        self.request_fid(rid, '1,2,3,16,9,8,10,13,14,15,16,4,267', '9999')

    def get_fid_array_tig_data_test(self, stock_code='229200', market_code='J', gid='1002', day1='', pnc='0', request_count=100):
        """
        :param stock_code:
        :param market_code:
        :param gid:
        :param day1: 가져오려는 마지막 틱 값으로, 출력 데이터 제일 처음에 위치한다.
        :param pnc:
        :param request_count: 틱 갯수
        :return:
        """
        rid = self.creat_request_id()
        assert rid, 'Request ID 가 생성되지 않았습니다.'
        self.set_fid_input_data(rid, '9001', market_code)
        self.set_fid_input_data(rid, '9002', stock_code)
        self.set_fid_input_data(rid, '9008', '4')
        self.set_fid_input_data(rid, '9034', day1)
        # self.set_fid_input_data(rid, '9035', day2)
        self.set_fid_input_data(rid, 'GID', gid)
        # 1. 종목코드
        # 3. 한글 종목명
        # 9. 일자 8자리
        # 8. 시간
        # 13.14.15.4. 시가 고가 저가 현재가
        # 10. 전날 종가
        # pnc = self.request_fid_array(rid, '1,3,9,8,13,14,15,4,10', '1', pnc, '9999', request_count)
        pnc = self.request_fid_array(rid, '9,8,4', '1', pnc, '9999', request_count)
        print('pnc', pnc)
        # row = self.get_fid_output_row_count(rid)
        # fid_output_data = self.get_fid_output_data(rid, pnc, row)

    def get_fid_array_tig_data_test_futures(self, stock_code='101P3000', market_code='F', gid='1002', day1='', pnc='1', request_count=100):
        """
        :param stock_code:
        :param market_code:
        :param gid:
        :param day1: 가져오려는 마지막 틱 값으로, 출력 데이터 제일 처음에 위치한다.
        :param pnc:
        :param request_count: 틱 갯수
        :return:
        """
        rid = self.creat_request_id()
        assert rid, 'Request ID 가 생성되지 않았습니다.'
        self.set_fid_input_data(rid, '9001', market_code)
        self.set_fid_input_data(rid, '9002', stock_code)
        self.set_fid_input_data(rid, '9034', day1)
        # self.set_fid_input_data(rid, '9035', day2)
        self.set_fid_input_data(rid, 'GID', gid)
        # 1. 종목코드
        # 3. 한글 종목명
        # 9. 일자 8자리
        # 8. 시간
        # 13.14.15.4. 시가 고가 저가 현재가
        # 10. 전날 종가
        # pnc = self.request_fid_array(rid, '1,3,9,8,13,14,15,4,10', '1', pnc, '9999', request_count)
        pnc = self.request_fid_array(rid, '9,8,4', '1', pnc, '9999', request_count)
        print('pnc', pnc)
        # row = self.get_fid_output_row_count(rid)
        # fid_output_data = self.get_fid_output_data(rid, pnc, row)

    # 실제 사용될 루프 코드
    def get_stock_tig_data(self, stock_code='229200', market_code='J', gid='1002', day1='', pnc='0', request_count=100, loop=True):
        """
        종목 코드와 날짜, 시장 구분자를 받아서 틱데이터를 가져오는 함수.

        :param stock_code: 주식 종목. 기본값인 '229200' 은 Kodex 코스닥 150
        :param market_code: 주식 시장의 경우 'J', 주식선물 시장의 경우 'JF', 지수선물 시장의 경우 'W'
        :param gid: 틱데이터를 가져올 때는 gid = '1002'
        :param day1: 가져오려는 마지막 틱 값의 날짜로, 출력 데이터 제일 처음에 위치한다.
            - 예를 들어, 20181217 이면, 2018년 12월 17일 틱 데이터 중 최신 값 부터 점점 오래된 값들을 불러온다.
        :param pnc: 중요하지 않은 값
        :param request_count: 가져오려는 틱 갯수. 갯수가 충분히 많으면 day1 이전 날짜의 틱 데이터도 가져온다.
        :return: None
        """
        rid = self.creat_request_id()
        assert rid, 'Request ID 가 생성되지 않았습니다.'

        self.set_fid_input_data(rid, '9001', market_code)
        self.set_fid_input_data(rid, '9002', stock_code)
        self.set_fid_input_data(rid, '9008', '4')
        self.set_fid_input_data(rid, '9034', day1)
        self.set_fid_input_data(rid, 'GID', gid)

        if loop:
            today = datetime.now()
            open_time = today.replace(hour=8, minute=55)
            close_time = today.replace(hour=15, minute=35)
            while open_time <= datetime.now() < close_time:
                self.get_stock_tig_data(loop=False)  # rid 및 기타 설정이 필요하므로
                time.sleep(1)  # 1초간 쉰다. 1초에 한번씩 가져오기 위한 것
        else:
            self.request_fid_array(rid, '9,8,4', '1', pnc, '9999', request_count)

    def get_futures_tig_data(self, code='101P3000', market_code='F', gid='1002', day1='', pnc='1', request_count=100, loop=True):
        """
        :param code: 선물 종목 코드. 기본값인 '101P3000'은 코스피 19년 3월물
        :param market_code: 선물 시장 타입. 지수 선물은 F, 주식 선물은 JF
        :param gid: 틱 데이터를 수신할 때는 1002 고정
        :param day1: 가져오려는 마지막 틱 값의 날짜로, 출력 데이터 제일 처음에 위치한다.
        :param pnc: 중요하지 않은 값
        :param request_count: 가져오려는 틱 갯수
        :return: None
        """
        rid = self.creat_request_id()
        assert rid, 'Request ID 가 생성되지 않았습니다.'

        self.set_fid_input_data(rid, '9001', market_code)
        self.set_fid_input_data(rid, '9002', code)
        self.set_fid_input_data(rid, '9034', day1)
        self.set_fid_input_data(rid, 'GID', gid)

        if loop:
            today = datetime.now()
            open_time = today.replace(hour=8, minute=55)
            close_time = today.replace(hour=15, minute=35)
            while open_time <= datetime.now() < close_time:
                self.get_futures_tig_data(loop=False)  # rid 및 기타 설정이 필요하므로
                time.sleep(1)  # 1초간 쉰다. 1초에 한번씩 가져오기 위한 것
        else:
            self.request_fid_array(rid, '9,8,4', '1', pnc, '9999', request_count)

    def get_real_data_test(self, real_name='S00', real_key='229200'):
        items = [
            'SHRN_ISCD',  # 종목코드
            'STCK_CNTG_HOUR',  # 체결시간
            'HOUR_CLS_CODE',  # 시간 구분 코드
            'STCK_OPRC',  # 시가
            'STCK_HGPR',  # 고가
            'STCK_LWPR',  # 저가
            'STCK_PRPR',  # 현재가
        ]
        self.register_real(real_name, real_key)
        self.get_real_output_data(real_name, items[-1])

    @property
    def user_id(self):
        return self._user_id

    @property
    def user_pw(self):
        return self._user_pw

    @property
    def cert_pw(self):
        return self._cert_pw

    @property
    def prog_id(self):
        return self.__class__._PROG_ID

    def set_user_id(self, user_id):
        self._user_id = encrypt(ENCRYPTION_KEY, user_id)

    def set_user_pw(self, user_pw):
        self._user_pw = encrypt(ENCRYPTION_KEY, user_pw)

    def set_cert_pw(self, cert_pw):
        self._cert_pw = encrypt(ENCRYPTION_KEY, cert_pw)

    def login(self, simulation=False, international=False):
        init_result = self.comm_init()
        assert init_result, '통신 모듈 초기화 과정에서 오류가 발생했습니다'

        if simulation:
            if international:
                self.set_login_mode(0, 2)
            else:
                self.set_login_mode(0, 1)

        login_result = self.comm_login()
        assert login_result, '로그인 과정에서 오류가 발생했습니다.'

        return self.get_login_state()

    def logout(self):
        self.comm_logout()

    def terminate(self, option=1):
        self.comm_terminate(option)

    # 통신 관련 메서드
    def comm_init(self):
        """
        통신 모듈 초기화 및 연결 함수. 로그인 처리 전에 호출한다.
        :return: 성공할 경우 True, 실패할 경우 False
        """
        result = self.dynamicCall('CommInit()')
        return True if result == 0 else False

    def comm_terminate(self, option=1):
        """
        통신 모듈 연결 해제. 로그아웃 처리 이후에 호출. 로그아웃을 하지 않고 호출할 경우에도 잘 작동함
        :param option: 1 = 통신 모듈 종료 및 연결 해제, 2 = 연결만 해제
        :return: None
        """
        return self.dynamicCall('CommTerminate(option)', option)

    def comm_get_connect_state(self):
        """
        통신 연결 상태 확인. comm_init() 메서드 호출 후 통신 연결 상태 확인을 위해 호출한다.
        :return: 연결된 경우 True, 연결이 끊긴 경우 False
        """
        result = self.dynamicCall('CommGetConnectState()')
        return True if result else False

    def comm_login(self):
        """
        로그인 처리 함수. comm_init() 호출 후 통신 연결이 완료된 뒤에 호출
        :return: 성공할 경우 True, 실패할 경우 False
        """
        user_id = decrypt(ENCRYPTION_KEY, self.user_id)
        user_pw = decrypt(ENCRYPTION_KEY, self.user_pw)
        cert_pw = decrypt(ENCRYPTION_KEY, self.cert_pw)
        result = self.dynamicCall('CommLogin(user_id, user_pw, cert_pw)', user_id, user_pw, cert_pw)
        return True if result else False

    def comm_logout(self):
        """
        로그아웃 함수. comm_terminate() 전에 호출한다.
        :return: 성공할 경우 True, 실패할 경우 False
        """
        user_id = decrypt(ENCRYPTION_KEY, self.user_id)
        result = self.dynamicCall('CommLogout(user_id)', user_id)
        return True if result == 0 else False

    def get_login_state(self):
        """
        로그인 상태 확인. comm_login() 메서드 호출 후 사용
        :return: 로그인 상태는 True, 로그아웃 상태는 False
        """
        result = self.dynamicCall('GetLoginState()')
        return True if result else False

    def set_login_mode(self, option, mode):
        """
        로그인 모드를 설정하는 함수. comm_login() 메서드 호출 전 설정.
            - (0, 0): 실거래 로그인
            - (0, 1): 국내 모의 로그인
            - (0, 2): 해외 모의 로그인
            - (1, 1): 시세 전용 로그인
        :param option: 투자 여부 설정
            - 0: 실거래 및 모의 투자 (*모의투자는 모의투자 전용 ID 가 필요하다)
            - 1: 시세 전용
        :param mode: 인증 모드 설정
            if option == 0:
                - 0: 실거래
                - 1: 국내 모의
                - 2: 해외 모의
            if option == 1:
                - 0: 공인인증
                - 1: 시세 전용 (공인인증 없음)
        :return: None
        """
        self.dynamicCall('SetLoginMode(option, mode)', option, mode)

    def get_login_mode(self, option=2):
        """
        로그인 상태를 확인해준다. comm_login() 메서드 호출 이후 로그인 상태 확인 목적으로 호출한다.
        :param option:
            - 0: 모의투자 여부 체크
            - 1: 시세 전용 여부 체크
            - 2: 직원/고객 로그인 체크
        :return:
            - -1: 실패
            - -1 보다 큰 정수: 성공
        """
        return self.dynamicCall('GetLoginMode(option)', option)

    # 리소스 관련 메서드
    # 하나대투에선 Tran(Transaction) 이라는 용어를 사용하는데, 따로 정의가 되어있지는 않으나,
    # '서버로부터 데이터를 주고 받는 행위' 정도로 생각하면 쉽다.
    def load_tran_resource(self, res_path):
        """
        기능: Tran 조회 I/O Block 정보 리소스 파일 로드
        호출: Tran 조회시에 반드시 리소스 파일이 에이전트 컨트롤에 적재되어 있어야 한다.
        :param res_path: 리소스 파일(*.res) 경로
        :return: 성공할 경우 True, 실패할 경우 False
        """
        return True if self.dynamicCall('LoadTranResource(res_path)', res_path) else False

    def load_real_resource(self, res_path):
        """
        기능: 실시간 Block 정보 리소스 파일 로드
        호출: 실시간 등록시에 반드시 리소스 파일이 에이전트 컨트롤에 적재되어 있어야 한다.
        :param res_path: 리소스 파일(*.res) 경로
        :return: 성공할 경우 True, 실패할 경우 False
        """
        return True if self.dynamicCall('LoadRealResource(res_path)', res_path) else False

    # 통신 조회 관련 공통
    def creat_request_id(self):
        """
        기능:
            - 조회 고유 ID 생성 (Request ID)
            - 호출할 때마다 '2' 부터 차례대로 크기가 증가하며 호출 된다.
            - 예를 들어 첫 번째 호출에선 2, 두 번째 호출에선 3, 세 번째 호출에선 4 ...
            - release_rq_id(2) 를 호출하면 값이 2 인 Request ID 가 할당 해제된다.
        호출: Tran/FID 조회시 Request ID 를 먼저 생성한다.
        :return: 신규 Request ID
        """
        return self.dynamicCall('CreateRequestID()')

    def get_comm_recv_option_value(self, option):
        """
        기능: 조회 응답 부가정보/옵션값 반환
        호출: Tran/FID 조회(on_get_tran_data(), on_get_fid_data()) 응답 이벤트 안에서만 호출한다.
            그렇지 않을 경우 빈 문자열 반환
        :param option:
            - 0: Tran code
            - 1: 연속데이터 구분 (return= 0: 없음, 1: 이전, 2: 다음, 3: 이전/다음)
            - 2: 연속조회키
            - 3: 메세지 코드
            - 4: 메세지
            - 5: 부가메세지 코드
            - 6: 부가메세지
        :return: option 에 대응하는 문자열
        """
        return self.dynamicCall('GetCommRecvOptionValue(option)', option)

    def release_rq_id(self, request_id):
        """
        기능: 조회 고유 ID (Request ID) 할당 해제
        호출: create_request_id() 함수로 생성한 ID 를 할당해제.
            입력된 ID 가 할당되지 않은 것이라도 오류가 발생하지 않음
        :param request_id: Request ID
        :return: None
        """
        return self.dynamicCall('ReleaseRqId(request_id)', request_id)

    # Tran 조회 관련 메서드
    def set_tran_input_data(self, request_id, tran_code, rec_name, item, value):
        """
        기능: Tran 조회, 항목별 입력값을 입력한다.
        호출: request_tran() 호출 전에 통신 Input 데이터 입력 목적으로 호출한다.
        :param request_id: Request ID (create_request_id) 메서드로 생성
        :param tran_code: 서비스 Tran 코드 (Tran 리소스파일(*.res)의 'TR_CODE=' 항목)
        :param rec_name: Input Record 이름 (Tran 리소스파일(*.res)의 'REC_NAME' 항목)
        :param item: Input 항목명 (Tran 리소스파일(*.res)의 'ITEM=' 항목)
        :param value: Input 항목에 대응하는 입력값
        :return: 성공할 경우 True, 실패할 경우 False
        """
        result = self.dynamicCall('SetTranInputData(r, t, rec, i, v)', request_id, tran_code, rec_name, item, value)
        return True if result else False

    def request_tran(self, request_id, tran_code, is_benefit, prev_or_next, prev_next_key,
                     screen_no, is_order, request_count):
        """
        기능: Tran 조회 요청
        호출: 서버에 Tran 조회 요청 시 호출

        :param request_id: 조회 고유 ID (Request ID) - (create_request_id() 로 생성)
        :param tran_code: 서비스 Tr 코드(Tran 리소스파일(*.res)파일의 ' TR_CODE=' 항목)
        :param is_benefit: 수익계좌 여부 True/False
        :param prev_or_next: 연속조회 구분 (return= "0": 일반조회, "1": 연속조회 첫 조회, "2": 이전조회, "3": 다음조회)
        :param prev_next_key: 다음/이전 조회시 연속구분이 되는 키값 입력 (조회응답으로 내려 온다.)
        :param screen_no: 화면번호 (ex. "9999")
        :param is_order: 주문 여부 (True/False)
        :param request_count: 조회 응답으로 받을 최대 데이터 건수 (Max: 9999)
        :return: 성공여부 True/False
        """
        is_benefit = 'Y' if is_benefit else 'N'
        tran_type = 'U' if is_order else 'Q'  # 'U' 는 update 라고 한다. 관례적인 표현이라고.
        result = self.dynamicCall(
            'RequestTran(r, t, ib, pn, pnk, sn, io, rc)',
            request_id, tran_code, is_benefit, prev_or_next,
            prev_next_key, screen_no, tran_type, request_count,
        )
        return True if result > 0 else False

    def get_tran_output_row_count(self, tran_code, rec_name):
        """
        기능: Tran 조회 응답 데이터 건수 반환
        호출: Tran 조회 응답 이벤트 (on_get_tran_data()) 안에서만 호출한다.
        :param tran_code: 서비스 Tran 코드 (Tran 리소스파일(*.res)파일의 ' TR_CODE=' 항목)
        :param rec_name: Input 레코드명(Tran 리소스파일(*.res)파일의 ' REC_NAME=' 항목)
        :return: 데이터 수
        """
        return self.dynamicCall('GetTranOutputRowCnt(t, rn)', tran_code, rec_name)

    def get_tran_output_data(self, tran_code, rec_name, item, row):
        """
        기능: Tran 조회 항목별 응답데이터 반환
        호출: Tran 조회 응답 이벤트(on_get_tran_data()) 안에서만 호출한다.
        :param tran_code: 서비스 Tr 코드(Tran 리소스파일(*.res)파일의 ' TR_CODE=' 항목)
        :param rec_name: Input Record 이름 (Tran 리소스파일(*.res)의 'REC_NAME' 항목)
        :param item: Input 항목명 (Tran 리소스파일(*.res)의 'ITEM=' 항목)
        :param row: ???
        :return: BSTR ???
        """
        return self.dynamicCall('GetTranOutputData(t, rn, i, r)', tran_code, rec_name, item, row)

    # FID 조회 관련
    def set_fid_input_data(self, request_id, fid, value):
        """
        기능: FID 조회 시, 항목별 입력값 입력
        호출: request_fid() 또는 request_fid_array() 호출 전에 조회 Input 데이터 입력 목적으로 호출한다.
        :param request_id: 조회고유 ID(Request ID)
        :param fid: FID 번호(ex-> "9002")
        :param value: FID 번호에 대응하는 입력값 (ex-> "000660")
        :return: 성공여부 True/False
        """
        return True if self.dynamicCall('SetFidInputData(r, f, v)', request_id, fid, value) else False

    def request_fid(self, request_id, output_fid_list, screen_no):
        """
        기능: FID 조회 요청 - 응답데이터가 단건(single)
        호출: 서버에 FID 조회 요청 시 호출(응답으로 받을 데이터 단건일 경우에 사용)
        :param request_id: 조회고유 ID(Request ID)
        :param output_fid_list: 응답으로 받을 FID 번호들(ex-> "4,6,5,7,11,28,13,14,15")
        :param screen_no: 화면번호 (ex-> "9999")
        :return: 음수 : 실패, 1 : 성공 : 2보다 큰 정수 (확인후 변경)
        """
        response = self.dynamicCall('RequestFid(r, o, s)', request_id, output_fid_list, screen_no)
        self.event_connect_loop.exec_()
        print('FID:', response)
        return response

    def request_fid_array(self, request_id, output_fid_list, pre_next, pre_next_context, screen_no, request_count):
        """
        기능: FID 조회 요청 - 응답 데이터가 복수건(array)
        호출: 서버에 FID 조회 요청 시 호출(응답받을 데이터가 복수건인 경우에 사용)
        :param request_id: 조회고유 ID(Request ID)
        :param output_fid_list: 응답으로 받을 FID 번호들(ex-> "4,6,5,7,11,28,13,14,15")
            - 1. 종목코드
            - 3. 한글 종목명
            - 9. 일자 8자리
            - 8. 시간
            - 13.14.15.4. 시가 고가 저가 현재가
            - 10. 전날 종가
        :param pre_next: 연속조회 구분 ("0" :일반, "1" : 연속 첫 조회, "2" : 이전 조회, "3" : 다음 조회)
        :param pre_next_context: 조회 응답으로 받은 연속거래키
        :param screen_no: 화면변호(ex-> "9999")
        :param request_count: 조회 응답으로 받을 최대 데이터 건수(Max : 9999)
        :return: 음수 : 실패, 1 : 성공 : 2보다 큰 정수 (확인후 변경)
        """
        self.fid = self.dynamicCall(
            'RequestFidArray(r, ofl, pn, pnc, sn, rc)',
            request_id, output_fid_list, pre_next, pre_next_context, screen_no, request_count
        )

        self.event_connect_loop.exec_()
        return self.fid

    def get_fid_output_row_count(self, request_id):
        """
        기능: FID 조회 응답데이터 건수
        호출: FID 조회 응답 이벤트(on_get_fid_data()) 안에서만 호출한다.
        :param request_id: 조회 고유 ID(Request ID)
        :return: 0 : 데이터 없음, 0보다 큰 정수 : 데이터 건수
        """
        return self.dynamicCall('GetFidOutputRowCnt(r)', request_id)

    def get_fid_output_data(self, request_id, fid, row):
        """
        기능: FID 조회 항목별 응답 데이터 반환
        호출: FID 조회응답 이벤트(OnGetFidData) 안에서만 호출한다.
        :param request_id: 조회 고유 ID(Request ID)
        :param fid: 응답 받은 FID 번호(ex-> "4")
        :param row: 항목값이 위치한 행 인덱스
            - 단건(single): 0
            - 복수건(array): 해당 행의 인덱스 번호
        :return: FID 에 대응한 응답 데이터
        """
        return self.dynamicCall('GetFidOutputData(r, f, ro)', request_id, fid, row)

    def set_portfolio_fid_input_data(self, request_id, symbol_code, symbol_market):
        """
        기능: 관심종목형(Portfolio) FID 조회 시, 항목별 입력값 입력
        호출: RequestFid 또는 RequestFidArray 호출 전에 조회 Input 데이터 입력 목적으로 호출한다.
        :param request_id: 조회고유ID(Request ID)
        :param symbol_code: 종목코드
        :param symbol_market: 종목 시장코드
        :return: 0 : 실패, 1 : 성공
        """
        return self.dynamicCall('SetPortfolioFidInputData(r, sc, sm)', request_id, symbol_code, symbol_market)

    # 실시간 관련 메서드
    def register_real(self, real_name, real_key):
        """
        기능: 실시간 등록한다.
        호출: 로그인 처리가 완료된 이후 또는 Tran/FID 조회 응답 이벤트 안에서 호출한다.
        :param real_name: 실시간 등록할 실시간코드명 (실시간 리소스파일(*.res)파일의 ' REAL_NAME=' 항목(ex-> "S00"))
        :param real_key: 실시간 수신 시 데이터 구분키가 될 값(ex-> "000660" : SK하이닉스 종목코드)
        :return: 성공 여부 True/False
        """
        result = self.dynamicCall('RegisterReal(rn, rk)', real_name, real_key)
        return True if result == 0 else False

    def un_register_real(self, real_name, real_key):
        """
        기능: 실시간등록 해제한다.
        호출: register_real 함수 호출 이후에 호출한다.
        :param real_name: 실시간 등록할 실시간코드명 (실시간 리소스파일(*.res)파일의 ' REAL_NAME=' 항목(ex-> "S00"))
        :param real_key: 실시간 수신 시 데이터 구분키가 될 값(ex-> "000660" : SK하이닉스 종목코드)
        :return: 성공 여부 True/False
        """
        result = self.dynamicCall('UnRegisterReal(rn, rk)', real_name, real_key)
        return True if result == 1 else False

    def all_un_register_real(self):
        """
        기능: 모든 실시간등록 해제한다.
        호출: register_real 함수 호출 이후에 호출한다
        :return: 성공 여부 True/False
        """
        return True if self.dynamicCall('AllUnRegisterReal()') == 1 else False

    def get_real_output_data(self, real_name, item):
        """
        기능: 항목별 실시간 수신 데이터를 반환한다.
        호출: 실시간데이터 수신 이벤트(on_get_real_data()) 안에서만 호출한다.
        :param real_name: 실시간 등록할 실시간코드명 (실시간 리소스파일(*.res)파일의 ' REAL_NAME=' 항목(ex-> "S00"))
        :param item: 실시간 리소스파일(*.res)파일의 ' ITEM=' 항목(ex-> " SHRN_ISCD")
        :return: item 에 해당하는 데이터 값
        """
        response = self.dynamicCall('GetRealOutputData(rn, i)', real_name, item)
        self.event_connect_loop.exec_()
        return response

    # 부가적인 메서드
    def get_last_err_msg(self):
        """
        기능: 에러 메시지 확인
        호출: 마지막으로 호출된 API 메소드에서 에러가 발생했을 경우, 에러메시지 확인하기 위해 호출한다.
        :return: 에러메시지
        """
        return self.dynamicCall('GetLastErrMsg()')

    def get_api_agent_module_path(self):
        """
        기능: OpenAPI 에이전트 모듈 파일경로 반환
        호출: 에이전트 오브젝트 생성 이후에 호출
        :return: 파일경로
        """
        return self.dynamicCall('GetApiAgentModulePath()')

    def get_encrypt(self, plain_text):
        """
        기능: 평문을 암호화한다(계좌비밀번호 암호화 등에 사용된다.)
        :param plain_text: 평문
        :return: 암호화된 문자열
        """
        return self.dynamicCall('GetEncrpyt(p)', plain_text)

    def set_off_agent_message_box(self, option):
        """
        기능: 에이전트에 띄우는 메시박스를 막는다.
        호출: CommInit 함수 호출전 SetOffAgentMessageBox함수를 호출해야된다.
        :param option:
            - 0 : 에이전트 메시지박스 실행,
            - 1 : 에이전트 메시지박스 실행 안 함
        :return: None
        """
        return self.dynamicCall('SetOffAgentMessageBox(o)', option)

    # 아직 추가되지 않은 기능인 듯
    def get_optional_function(self, option, n_val1, n_val2, str_val1, str_val2):
        """
        기능: 부가적인 옵션 처리(옵션 세부 설정은 신규 추가 시 가이드 문서에 포함 재배포)
        :param option:
        :param n_val1:
        :param n_val2:
        :param str_val1:
        :param str_val2:
        :return: 옵션 처리 결과 문자열
        """
        return self.dynamicCall('SetOptionalFunction(o, n1, n2, s1, s2)', option, n_val1, n_val2, str_val1, str_val2)

    # 계좌 관련 메서드들
    def get_acc_info(self, option, acc_no):
        """
        기능: 계좌 정보
        :param option:
            - 0: 계좌대체번호
            - 1: 계좌상품번호
            - 198: 대리인 등록 여부 (return= 'Y', 주문대리 계좌일 경우)
        :param acc_no:
        :return:
        """
        return self.dynamicCall('GetAccInfo(0, a)', option, acc_no)

    def get_user_acc_count(self):
        """
        기능: 보유 계좌 수
        :return: 보유 계좌 수
        """
        return self.dynamicCall('GetUserAccCnt()')

    def get_user_acc_no(self, index):
        """
        기능: 보유 계좌 반환
        :param index: 보유 계좌 인덱스
        :return: 계좌 정보 반환 (종합계좌번호, 계좌 상품 번호)
        """
        return self.dynamicCall('GetUserAccNo(i)', index)


class HanaUser(Hana):
    def __init__(self, user_name=None):
        from config_secret import USER_INFO

        user_info = USER_INFO
        if user_name:
            user_info = USER_INFO[user_name]
        user_id = decrypt(ENCRYPTION_KEY, user_info['user_id'])
        user_pw = decrypt(ENCRYPTION_KEY, user_info['user_pw'])
        cert_pw = decrypt(ENCRYPTION_KEY, user_info['cert_pw'])

        super().__init__(user_id, user_pw, cert_pw)

        count = 1
        while not self.login():
            count += 1
            if count > 100:
                break
            print(f'login 에 실패하여 재요청 중입니다 ({count} 번째 요청 중)')


